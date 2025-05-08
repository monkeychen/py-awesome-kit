import os
import sys
from openpyxl import load_workbook


def restore_merged_cells(input_file, output_file):
    # 加载工作簿
    wb = load_workbook(input_file)
    
    # 遍历每个工作表
    for sheet in wb.worksheets:
        # 获取合并单元格列表
        merged_cells = sheet.merged_cells.ranges.copy()
        
        # 还原合并单元格
        for merged_cell in merged_cells:
            min_col, min_row, max_col, max_row = merged_cell.bounds
            
            # 获取合并单元格的原始值（仅左上角有效）
            original_value = sheet.cell(row=min_row, column=min_col).value
            
            # 先解除合并单元格
            sheet.unmerge_cells(str(merged_cell))
            
            # 将原合并区域所有单元格填充为原始值
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = original_value
    
    # 保存新文件
    wb.save(output_file)


def process_files_in_directory(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            input_file = os.path.join(directory, filename)
            output_file = os.path.join(directory, f'restored_{filename}')
            print(f'Processing {filename}...')
            restore_merged_cells(input_file, output_file)
            print(f'Saved as {output_file}')


if __name__ == '__main__':
    temp_directory = sys.argv[1]
    process_files_in_directory(temp_directory)
